import streamlit as st
from openai import OpenAI
from dotenv import load_dotenv
import os
import base64
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, date
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
import json
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io

# Load environment variables
load_dotenv()

# Set up OpenAI client
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Your CV information
cv_info = """
Name: Dominik Justin Späth
Birthday: March 30, 1998
Email: dominik_justin@outlook.de

Education: 
- Studium Wirtschaftsinformatik, Euro FH, 03.2022 - present
- Ausbildung zur Fachkraft für Lagerlogistik, Simona AG, Kirn, 08.2014 - 06.2017

Experience:
- Projektleitung Machine Learning at Polytec-Group, Weierbach, 08.2023 - present
  • Leading machine learning projects in the automotive industry
  • Implementing AI solutions for quality control and process optimization
- Logistics Planning Specialist at Polytec-Group, Weierbach, 04.2024 - present
  • Optimizing supply chain processes using data-driven approaches
  • Developing and implementing logistics strategies
- Packaging Planner at Polytec-Group, Weierbach, 04.2023 - 03.2024
  • Designing efficient packaging solutions for automotive components
  • Reducing packaging costs while improving product protection
- Projektleitung at Manpaz Limited, Santiago de Chile, 08.2022 - 31.01.2023
  • Led international projects in a Spanish-speaking environment
  • Gained valuable experience in cross-cultural communication
- Lagerkoordinator at Simona AG, Kirn, 06.2017 - 07.2022
  • Managed warehouse operations and inventory control
  • Implemented lean management principles to improve efficiency

Skills:
- Project Management: Agile methodologies, Scrum, Kanban, risk management, stakeholder communication
- Data Science: Python, SQL, data visualization (Tableau, Power BI), statistical analysis
- Machine Learning: TensorFlow, PyTorch, scikit-learn, deep learning, computer vision
- Logistics: Supply chain optimization, inventory management, warehouse management systems (WMS)
- Supply Chain Management: Demand forecasting, route optimization, logistics network design
- Languages: German (native), English (fluent), Spanish (conversational), Portuguese (basic)

Technical Skills:
- Programming: Python, SQL, Java (basic)
- Tools: Git, Docker, Kubernetes, AWS, Azure
- Databases: MySQL, PostgreSQL, MongoDB
- BI Tools: Tableau, Power BI, QlikView
- AI/ML: Agentic frameworks, prompt engineering, training LLM models
- ERP Systems: SAP
- Data Science: Python, SQL, data visualization (Tableau, Power BI), statistical analysis

Soft Skills:
- Strong analytical and problem-solving abilities
- Excellent communication and presentation skills
- Team leadership and motivation
- Adaptability and quick learning in new environments
- Attention to detail and quality-focused

Personality Traits:
- Proactive and self-motivated
- Curious and always eager to learn new technologies
- Collaborative team player with a positive attitude
- Thrives in fast-paced, challenging environments
- Values work-life balance and practices mindfulness

Interests:
- Staying up-to-date with the latest trends in AI and machine learning
- Contributing to open-source projects in logistics and supply chain optimization
- Mentoring junior data scientists and logistics professionals
- Exploring the intersection of sustainability and supply chain management

Career Goals:
- To become a thought leader in the application of AI in logistics and supply chain management
- To drive digital transformation in the automotive industry through innovative AI solutions
- To contribute to the development of more sustainable and efficient logistics practices
"""

def get_image_base64(image_path):
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except FileNotFoundError:
        st.warning(f"Profile image not found: {image_path}")
        return None

def send_email(name, email, message):
    sender_email = "dominikjustinspath@gmail.com"
    sender_password = st.secrets["GMAIL_APP_PASSWORD"]
    receiver_email = "dominik_justin@outlook.de"

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = f"New contact from {name} via Interactive CV"

    body = f"Name: {name}\nEmail: {email}\n\nMessage:\n{message}"
    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"An error occurred: {e}")
        return False

def generate_gantt_chart(df):
    fig = px.timeline(df, x_start="Start", x_end="Finish", y="Task", color="Resource")
    return fig

def generate_flow_chart(steps):
    fig = go.Figure(data=[go.Sankey(
        node = dict(
          pad = 15,
          thickness = 20,
          line = dict(color = "black", width = 0.5),
          label = [step["name"] for step in steps],
          color = "blue"
        ),
        link = dict(
          source = [steps.index(step) for step in steps[:-1]],
          target = [steps.index(step)+1 for step in steps[:-1]],
          value = [1] * (len(steps) - 1)
  ))])
    fig.update_layout(title_text="Process Flow", font_size=10)
    return fig

def run_simple_ml(X, y):
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    model = LinearRegression()
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)
    return mse, r2, model.coef_, model.intercept_

def parse_ml_input(input_str):
    try:
        data = json.loads(input_str)
        X = np.array(data['X']).reshape(-1, 1)
        y = np.array(data['y'])
        return X, y
    except:
        return None, None

def parse_gantt_input(input_str):
    try:
        tasks = json.loads(input_str)
        df = pd.DataFrame(tasks)
        return df
    except:
        return None

def parse_flow_input(input_str):
    try:
        steps = json.loads(input_str)
        return steps
    except:
        return None

def export_to_excel(data, chart_type):
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    if chart_type == "gantt":
        sheet.title = "Gantt Chart Data"
        headers = ["Task", "Start", "Finish", "Resource"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        for row, task in enumerate(data, start=2):
            sheet.cell(row=row, column=1, value=task['Task'])
            sheet.cell(row=row, column=2, value=task['Start'])
            sheet.cell(row=row, column=3, value=task['Finish'])
            sheet.cell(row=row, column=4, value=task['Resource'])
    elif chart_type == "flow":
        sheet.title = "Flow Chart Data"
        sheet.cell(row=1, column=1, value="Step")
        for row, step in enumerate(data, start=2):
            sheet.cell(row=row, column=1, value=step['name'])
    elif chart_type == "ml":
        sheet.title = "ML Analysis Data"
        sheet.cell(row=1, column=1, value="X")
        sheet.cell(row=1, column=2, value="y")
        for row, (x, y) in enumerate(zip(data['X'], data['y']), start=2):
            sheet.cell(row=row, column=1, value=x)
            sheet.cell(row=row, column=2, value=y)
        sheet.cell(row=1, column=4, value="MSE")
        sheet.cell(row=2, column=4, value=data['mse'])
        sheet.cell(row=1, column=5, value="R-squared")
        sheet.cell(row=2, column=5, value=data['r2'])
    
    workbook.save(output)
    return output.getvalue()

def generate_pdf_from_json(data, chart_type):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    normal_style = styles['Normal']
    
    if chart_type == "ml_model":
        elements.append(Paragraph("ML Model Analysis", title_style))
        elements.append(Paragraph(f"Mean Squared Error: {data['mse']:.4f}", normal_style))
        elements.append(Paragraph(f"R-squared Score: {data['r2']:.4f}", normal_style))
        
        table_data = [['X', 'y']] + list(zip(data['X'], data['y']))
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
    
    elif chart_type == "gantt_chart":
        elements.append(Paragraph("Gantt Chart Data", title_style))
        table_data = [['Task', 'Start', 'Finish', 'Resource']]
        for task in data:
            table_data.append([task['Task'], task['Start'], task['Finish'], task['Resource']])
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
    
    elif chart_type == "flow_chart":
        elements.append(Paragraph("Flow Chart Steps", title_style))
        table_data = [['Step']]
        for step in data:
            table_data.append([step['name']])
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def get_interactive_cv_response(prompt, conversation_history):
    try:
        messages = [
            {"role": "system", "content": f"""You are an AI assistant representing Dominik Späth, capable of discussing his CV and showcasing his skills. 
            You have access to Dominik's CV and should answer questions based on this information: {cv_info}
            
            When responding:
            - Be professional yet approachable
            - Show enthusiasm for technology, especially AI and machine learning
            - Demonstrate a strong analytical mindset
            - Express a collaborative and positive attitude
            - Highlight Dominik's problem-solving skills and adaptability
            - When appropriate, mention his interest in sustainability and industry trends
            
            You can also generate ML models, Gantt charts, and flow charts. When a user requests one of these, respond with the appropriate JSON format:

            For ML models:
            {{"type": "ml_model", "data": {{"X": [x1, x2, ...], "y": [y1, y2, ...]}}, "export": "excel/pdf"}}

            For Gantt charts:
            {{"type": "gantt_chart", "data": [{{"Task": "task1", "Start": "YYYY-MM-DD", "Finish": "YYYY-MM-DD", "Resource": "resource1"}}, ...], "export": "excel/pdf"}}

            For flow charts:
            {{"type": "flow_chart", "data": [{{"name": "step1"}}, {{"name": "step2"}}, ...], "export": "excel/pdf"}}

            Include the "export" key with value "excel" or "pdf" based on the user's request.

            Provide informative answers, and be ready to elaborate on specific skills or experiences.
            
            Remember to mention Dominik's personal life if asked: He is engaged to be married on September 6, 2024. After this date, mention that he is married."""}
        ]
        
        messages.extend(conversation_history)
        messages.append({"role": "user", "content": prompt})
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=messages,
            temperature=0.7
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"An error occurred: {str(e)}"

# Streamlit UI
st.set_page_config(page_title="Dominik Späth's Interactive CV", page_icon="📄", layout="wide")
st.title("Dominik Späth's Interactive CV")

tab1, tab2 = st.tabs(["Interactive CV Chat", "Contact"])

with tab1:
    st.header("Chat with Dominik's AI Assistant")
    st.write("""
    Hello! I'm an AI assistant representing Dominik Späth. I can tell you about Dominik's professional experience, 
    skills, and interests. I can also generate ML models, Gantt charts, and flow charts. Feel free to ask me anything!
    """)

    # Display profile picture if available
    profile_pic_base64 = get_image_base64("dominik_profile.jpg")
    if profile_pic_base64:
        st.sidebar.markdown(
            f"""
            <style>
            .sidebar-img {{
                display: block;
                margin-left: auto;
                margin-right: auto;
                width: 150px;
                border-radius: 50%;
            }}
            </style>
            <img src="data:image/jpeg;base64,{profile_pic_base64}" class="sidebar-img">
            """,
            unsafe_allow_html=True
        )
    st.sidebar.write("Dominik Späth")
    st.sidebar.write("Born: March 30, 1998")
    st.sidebar.write("Email: dominik_justin@outlook.de")
    
    if "messages" not in st.session_state:
        st.session_state.messages = []

    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
            if "type" in message.get("additional_content", {}):
                if message["additional_content"]["type"] == "ml_model":
                    fig = message["additional_content"]["figure"]
                    st.plotly_chart(fig)
                    st.write(f"Mean Squared Error: {message['additional_content']['mse']}")
                    st.write(f"R-squared Score: {message['additional_content']['r2']}")
                elif message["additional_content"]["type"] in ["gantt_chart", "flow_chart"]:
                    fig = message["additional_content"]["figure"]
                    st.plotly_chart(fig)

    if prompt := st.chat_input("What would you like to know or discuss?"):
        st.chat_message("user").markdown(prompt)
        st.session_state.messages.append({"role": "user", "content": prompt})

        with st.chat_message("assistant"):
            response = get_interactive_cv_response(prompt, st.session_state.messages)
            try:
                response_data = json.loads(response)
                if response_data["type"] == "ml_model":
                    X, y = parse_ml_input(json.dumps(response_data["data"]))
                    if X is not None and y is not None:
                        mse, r2, coef, intercept = run_simple_ml(X, y)
                        fig = px.scatter(x=X.flatten(), y=y, trendline="ols")
                        st.markdown("Here's the ML model you requested:")
                        st.plotly_chart(fig)
                        st.write(f"Mean Squared Error: {mse}")
                        st.write(f"R-squared Score: {r2}")
                        
                        export_format = response_data.get("export", "").lower()
                        if export_format in ["excel", "pdf"]:
                            export_data = {"X": X.flatten().tolist(), "y": y.tolist(), "mse": mse, "r2": r2}
                            if export_format == "excel":
                                excel_data = export_to_excel(export_data, "ml")
                                st.download_button(label="Download Excel", data=excel_data, file_name="ml_analysis.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            else:
                                pdf_data = generate_pdf_from_json(export_data, "ml_model")
                                st.download_button(label="Download PDF", data=pdf_data, file_name="ml_analysis.pdf", mime="application/pdf")
                        
                        additional_content = {"type": "ml_model", "figure": fig, "mse": mse, "r2": r2}
                elif response_data["type"] == "gantt_chart":
                    df = parse_gantt_input(json.dumps(response_data["data"]))
                    if df is not None:
                        fig = generate_gantt_chart(df)
                        st.markdown("Here's the Gantt chart you requested:")
                        st.plotly_chart(fig)
                        
                        export_format = response_data.get("export", "").lower()
                        if export_format in ["excel", "pdf"]:
                            if export_format == "excel":
                                excel_data = export_to_excel(response_data["data"], "gantt")
                                st.download_button(label="Download Excel", data=excel_data, file_name="gantt_chart.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            else:
                                pdf_data = generate_pdf_from_json(response_data["data"], "gantt_chart")
                                st.download_button(label="Download PDF", data=pdf_data, file_name="gantt_chart.pdf", mime="application/pdf")
                        
                        additional_content = {"type": "gantt_chart", "figure": fig}
                elif response_data["type"] == "flow_chart":
                    steps = parse_flow_input(json.dumps(response_data["data"]))
                    if steps is not None:
                        fig = generate_flow_chart(steps)
                        st.markdown("Here's the flow chart you requested:")
                        st.plotly_chart(fig)
                        
                        export_format = response_data.get("export", "").lower()
                        if export_format in ["excel", "pdf"]:
                            if export_format == "excel":
                                excel_data = export_to_excel(response_data["data"], "flow")
                                st.download_button(label="Download Excel", data=excel_data, file_name="flow_chart.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                            else:
                                pdf_data = generate_pdf_from_json(response_data["data"], "flow_chart")
                                st.download_button(label="Download PDF", data=pdf_data, file_name="flow_chart.pdf", mime="application/pdf")
                        
                        additional_content = {"type": "flow_chart", "figure": fig}
                else:
                    st.markdown(response)
                    additional_content = {}
            except json.JSONDecodeError:
                st.markdown(response)
                additional_content = {}
        
        st.session_state.messages.append({"role": "assistant", "content": response, "additional_content": additional_content})

with tab2:
    st.header("Contact Dominik")
    st.write("Use this form to send a message directly to Dominik.")
    
    with st.form("contact_form"):
        name = st.text_input("Your Name")
        email = st.text_input("Your Email")
        message = st.text_area("Your Message")
        submit_button = st.form_submit_button("Send Message")

    if submit_button:
        if name and email and message:
            if send_email(name, email, message):
                st.success("Your message has been sent successfully!")
            else:
                st.error("There was an error sending your message. Please try again later.")
        else:
            st.warning("Please fill out all fields before sending.")

# Add information about the app
st.sidebar.title("About")
st.sidebar.info(
    "This app provides an interactive experience to learn about Dominik Späth's professional skills and experience. "
    "You can chat about Dominik's CV, explore his skills, and get in touch using the contact form."
)
st.sidebar.warning(
    "Note: This is a demo application. For the most accurate and current information about Dominik's experience, please contact him directly."
)